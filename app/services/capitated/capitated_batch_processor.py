"""
Procesador de lotes de carga mensual de asegurados capitados.

Migrado de App\\Services\\Capitated\\CapitatedBatchProcessor.php.

NOTA: Este servicio requiere ``openpyxl`` para la lectura de archivos Excel.
      Verificar que esté instalado: pip install openpyxl
      (No se encontro en las dependencias del proyecto al momento de la migracion.)
"""

from __future__ import annotations

import hashlib
import json
import logging
import re
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from fastapi import UploadFile
from sqlalchemy import select, text, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.models.capitated_batch_item_log import CapitatedBatchItemLog
from app.models.capitated_batch_log import CapitatedBatchLog
from app.models.capitated_contract import CapitatedContract
from app.models.capitated_monthly_record import CapitatedMonthlyRecord
from app.models.capitated_product_insured import CapitatedProductInsured
from app.models.company import Company
from app.models.country import Country
from app.models.plan_version import PlanVersion, plan_version_countries, plan_version_repatriation_countries
from app.models.product import Product
from app.services.uploaded_file.uploaded_file_service import UploadedFileService
from app.support.capitated_rejection_codes import CapitatedRejectionCodes

logger = logging.getLogger(__name__)

SOURCE_EXCEL = "excel"


class CapitatedBatchProcessor:
    """Procesa un Excel de carga mensual y devuelve el batch generado."""

    def __init__(self, db: AsyncSession) -> None:
        self._db = db
        self._uploaded_file_service = UploadedFileService(db)

    # ------------------------------------------------------------------
    # processExcel  (entry point)
    # ------------------------------------------------------------------

    async def process_excel(
        self,
        company: Company,
        uploaded_file: UploadFile,
        coverage_month: date,
        is_any_month_allowed: bool,
        cutoff_day: int | None,
        user_id: int,
    ) -> CapitatedBatchLog:
        """
        Procesa un Excel de carga mensual y devuelve el CapitatedBatchLog resultante.

        No lanza excepciones hacia el controlador: siempre devuelve un batch
        con status draft/processed/failed y campos de resumen rellenados.
        """
        coverage_month = self._normalize_coverage_month(coverage_month)

        original_filename = uploaded_file.filename or "file.xlsx"

        # Validacion basica del archivo
        if not uploaded_file.filename:
            batch = CapitatedBatchLog(
                company_id=company.id,
                coverage_month=coverage_month,
                source=SOURCE_EXCEL,
                source_file_id=None,
                original_filename=original_filename,
                file_hash=None,
                created_by_user_id=user_id,
                status=CapitatedBatchLog.STATUS_FAILED,
                is_any_month_allowed=is_any_month_allowed,
                cutoff_day=cutoff_day,
                processed_at=datetime.now(timezone.utc),
                summary_json=json.dumps(
                    {"error_summary": "El archivo subido no es valido."},
                    ensure_ascii=False,
                ),
            )
            self._db.add(batch)
            await self._db.flush()
            return batch

        # 1) Crear batch en estado draft
        batch = CapitatedBatchLog(
            company_id=company.id,
            coverage_month=coverage_month,
            source=SOURCE_EXCEL,
            source_file_id=None,
            original_filename=original_filename,
            file_hash=None,
            created_by_user_id=user_id,
            status=CapitatedBatchLog.STATUS_DRAFT,
            is_any_month_allowed=is_any_month_allowed,
            cutoff_day=cutoff_day,
        )
        self._db.add(batch)
        await self._db.flush()

        # 2) Guardar archivo usando UploadedFileService
        storage_path = f"capitados/batches/{batch.id}"
        file_model = await self._uploaded_file_service.store(
            uploaded_file,
            uploaded_by_user_id=user_id,
            base_path=storage_path,
        )

        local_path = self._uploaded_file_service.get_local_path(file_model)
        file_hash = self._sha1_file(local_path)

        batch.source_file_id = file_model.id
        batch.file_hash = file_hash
        await self._db.flush()

        # 3) Intentar leer el Excel con openpyxl
        try:
            import openpyxl  # noqa: F811
        except ImportError:
            batch.status = CapitatedBatchLog.STATUS_FAILED
            batch.processed_at = datetime.now(timezone.utc)
            batch.summary_json = json.dumps(
                {"error_summary": "openpyxl no esta instalado en el servidor."},
                ensure_ascii=False,
            )
            await self._db.flush()
            return batch

        try:
            wb = openpyxl.load_workbook(local_path, read_only=True, data_only=True)
        except Exception as exc:
            batch.status = CapitatedBatchLog.STATUS_FAILED
            batch.processed_at = datetime.now(timezone.utc)
            batch.summary_json = json.dumps(
                {"error_summary": f"No se pudo leer el archivo Excel: {str(exc)[:500]}"},
                ensure_ascii=False,
            )
            await self._db.flush()
            return batch

        # 4) Procesar filas dentro de transaccion (el caller controla el commit/rollback)
        try:
            await self._process_spreadsheet_rows(batch, company, wb, coverage_month)
            await self._refresh_contracts_status(company)
            await self._db.flush()
            return batch
        except Exception as exc:
            if batch.status != CapitatedBatchLog.STATUS_FAILED:
                batch.status = CapitatedBatchLog.STATUS_FAILED
                batch.processed_at = datetime.now(timezone.utc)
                batch.summary_json = json.dumps(
                    {"error_summary": str(exc)[:1000]},
                    ensure_ascii=False,
                )
            await self._db.flush()
            return batch
        finally:
            wb.close()

    # ------------------------------------------------------------------
    # processSpreadsheetRows
    # ------------------------------------------------------------------

    async def _process_spreadsheet_rows(
        self,
        batch: CapitatedBatchLog,
        company: Company,
        wb: Any,
        coverage_month: date,
    ) -> None:
        """Logica principal fila a fila."""
        # 1) Validar estructura por hoja/producto/plan_version
        sheet_metas, plan_errors = await self._build_sheet_metas(company, wb)

        if plan_errors:
            batch.status = CapitatedBatchLog.STATUS_FAILED
            batch.total_plan_errors = len(plan_errors)
            batch.total_rows = 0
            batch.total_applied = 0
            batch.total_rejected = 0
            batch.total_duplicated = 0
            batch.total_incongruences = 0
            batch.total_rolled_back = 0
            batch.processed_at = datetime.now(timezone.utc)
            batch.summary_json = json.dumps(
                {
                    "error_summary": "Errores de estructura/plan impiden procesar el archivo.",
                    "plan_errors": plan_errors,
                },
                ensure_ascii=False,
            )
            raise RuntimeError("Errores de estructura/plan en el Excel de capitados.")

        # 2) Flags para logging de items
        log_applied = settings.capitados_batch_item_log_applied
        log_rejected = settings.capitados_batch_item_log_rejected
        log_incongruence = settings.capitados_batch_item_log_incongruence
        log_duplicated = settings.capitados_batch_item_log_duplicated

        # 3) Contadores
        total_rows = 0
        total_applied = 0
        total_rejected = 0
        total_duplicated = 0
        total_incongruences = 0
        errors_by_code: dict[str, int] = {}

        for sheet_name, meta in sheet_metas.items():
            sheet = meta["sheet"]
            product: Product = meta["product"]
            plan_version: PlanVersion = meta["plan_version"]
            columns: dict[str, int] = meta["columns"]
            res_allowed: dict[str, bool] = meta["residence_iso3"]
            rep_allowed: dict[str, bool] = meta["repatriation_iso3"]
            country_prices: dict[str, float] = meta["country_prices"]
            age_surcharges = meta["age_surcharges"]

            for row_idx, row_data in enumerate(
                sheet.iter_rows(min_row=2, values_only=False), start=2
            ):
                # Leer valores de la fila por indice de columna (1-based en openpyxl)
                values = {
                    "ID": self._cell_str(row_data, columns["ID"]),
                    "NOMBRE": self._cell_str(row_data, columns["NOMBRE"]),
                    "RESIDENCIA": self._cell_str(row_data, columns["RESIDENCIA"]),
                    "NACIONALIDAD": self._cell_str(row_data, columns["NACIONALIDAD"]),
                    "SEXO": self._cell_str(row_data, columns["SEXO"]),
                    "EDAD": self._cell_str(row_data, columns["EDAD"]),
                }

                # Fila totalmente vacia -> break
                if self._is_row_empty(values):
                    break

                total_rows += 1

                result = "applied"
                rejection_code: str | None = None
                rejection_detail: str | None = None

                person: CapitatedProductInsured | None = None
                contract: CapitatedContract | None = None
                monthly_record: CapitatedMonthlyRecord | None = None
                duplicated_record: CapitatedMonthlyRecord | None = None

                document_number = values["ID"].strip()
                full_name = values["NOMBRE"].strip()
                sex_raw = values["SEXO"].strip().upper()[:1]
                age_raw = values["EDAD"].strip()
                res_raw = values["RESIDENCIA"]
                rep_raw = values["NACIONALIDAD"]

                # Validaciones basicas
                if document_number == "":
                    result = "rejected"
                    rejection_code = CapitatedRejectionCodes.UNKNOWN_ERROR
                    rejection_detail = "El campo ID (numero de documento) es obligatorio."
                elif full_name == "":
                    result = "rejected"
                    rejection_code = CapitatedRejectionCodes.UNKNOWN_ERROR
                    rejection_detail = "El campo Nombre es obligatorio."
                elif sex_raw not in ("M", "F"):
                    result = "rejected"
                    rejection_code = CapitatedRejectionCodes.PERSON_SEX_INVALID
                    rejection_detail = "Sexo invalido, se espera M o F."

                sex = sex_raw
                age: int | None = None

                # Edad
                if result == "applied":
                    if age_raw == "" or not age_raw.replace(".", "", 1).lstrip("-").isdigit():
                        result = "rejected"
                        rejection_code = CapitatedRejectionCodes.PERSON_AGE_INVALID
                        rejection_detail = "Edad invalida o vacia."
                    else:
                        age = int(float(age_raw))
                        if age < 0:
                            result = "rejected"
                            rejection_code = CapitatedRejectionCodes.PERSON_AGE_INVALID
                            rejection_detail = "La edad debe ser mayor o igual a 0."

                # Paises
                res_iso3: str | None = None
                rep_iso3: str | None = None
                res_country: Country | None = None
                rep_country: Country | None = None

                if result == "applied":
                    res_iso3 = await self._normalize_country_to_iso3(res_raw)
                    if not res_iso3:
                        result = "rejected"
                        rejection_code = CapitatedRejectionCodes.PERSON_COUNTRY_CODE_NOT_FOUND
                        rejection_detail = "No se pudo resolver pais de residencia a ISO3."
                    else:
                        res_country = await self._find_country_by_iso3(res_iso3)
                        if not res_country:
                            result = "rejected"
                            rejection_code = CapitatedRejectionCodes.PERSON_COUNTRY_CODE_NOT_FOUND
                            rejection_detail = "No se pudo resolver pais de residencia a registro de pais."

                if result == "applied":
                    rep_iso3 = await self._normalize_country_to_iso3(rep_raw)
                    if not rep_iso3:
                        result = "rejected"
                        rejection_code = CapitatedRejectionCodes.PERSON_COUNTRY_CODE_NOT_FOUND
                        rejection_detail = "No se pudo resolver pais de repatriacion a ISO3."
                    else:
                        rep_country = await self._find_country_by_iso3(rep_iso3)
                        if not rep_country:
                            result = "rejected"
                            rejection_code = CapitatedRejectionCodes.PERSON_COUNTRY_CODE_NOT_FOUND
                            rejection_detail = "No se pudo resolver pais de repatriacion a registro de pais."

                if result == "applied":
                    if res_iso3 not in res_allowed:
                        result = "rejected"
                        rejection_code = CapitatedRejectionCodes.PERSON_RESIDENCE_NOT_ALLOWED
                        rejection_detail = "Pais de residencia no permitido para esta version de plan."

                if result == "applied" and rep_allowed:
                    if rep_iso3 not in rep_allowed:
                        result = "rejected"
                        rejection_code = CapitatedRejectionCodes.PERSON_REPATRIATION_NOT_ALLOWED
                        rejection_detail = "Pais de repatriacion no permitido para esta version de plan."

                # Persona: buscar/crear y detectar incongruencia
                if result in ("applied", "incongruence"):
                    person = await self._find_person(
                        company.id, product.id, document_number
                    )

                    if person:
                        existing_name = (person.full_name or "").strip()
                        existing_sex = (person.sex or "").upper()

                        if (
                            (existing_name != "" and existing_name != full_name)
                            or (existing_sex != "" and existing_sex != sex)
                        ):
                            result = "incongruence"
                            rejection_code = CapitatedRejectionCodes.PERSON_INCONGRUENCE
                            rejection_detail = (
                                "Datos incongruentes con la ficha existente de la persona (nombre/sexo)."
                            )
                        elif result == "applied":
                            person.full_name = full_name
                            person.sex = sex
                            person.residence_country_id = res_country.id if res_country else None
                            person.repatriation_country_id = rep_country.id if rep_country else None
                            person.age_reported = age
                            await self._db.flush()
                    else:
                        if result == "applied":
                            person = CapitatedProductInsured(
                                company_id=company.id,
                                product_id=product.id,
                                document_number=document_number,
                                full_name=full_name,
                                sex=sex,
                                residence_country_id=res_country.id if res_country else None,
                                repatriation_country_id=rep_country.id if rep_country else None,
                                age_reported=age,
                                status=CapitatedProductInsured.STATUS_ACTIVE,
                            )
                            self._db.add(person)
                            await self._db.flush()

                # Contrato y registro mensual
                if result == "applied" and person:
                    # 1) Duplicado exacto de mes
                    existing_record = await self._find_active_monthly_record(
                        company.id, product.id, person.id, coverage_month
                    )

                    if existing_record:
                        result = "duplicated"
                        rejection_code = CapitatedRejectionCodes.PERSON_DUPLICATED
                        rejection_detail = "Ya existe un registro aprobado para esta persona/producto y mes."
                        duplicated_record = existing_record
                    else:
                        # 2) Continuidad simple
                        last_record = await self._find_last_active_record(
                            company.id, product.id, person.id
                        )

                        contract = None
                        continuity_break = False

                        if last_record:
                            last_month = last_record.coverage_month.replace(day=1)

                            # Retroactivo
                            if coverage_month < last_month:
                                result = "rejected"
                                rejection_code = CapitatedRejectionCodes.RETROACTIVE_NOT_ALLOWED
                                rejection_detail = "No se permite cargar retroactivo anterior al ultimo mes aprobado."
                            else:
                                existing_contract = await self._find_contract_for_record(
                                    last_record.contract_id
                                )

                                # Calculo del mes esperado
                                expected_next = _add_month(last_month)

                                if coverage_month > expected_next:
                                    # Quiebre de continuidad
                                    if plan_version.max_entry_age and age is not None:
                                        if age > plan_version.max_entry_age:
                                            result = "rejected"
                                            rejection_code = CapitatedRejectionCodes.PERSON_AGE_INVALID
                                            rejection_detail = "Excede edad maxima para contratacion"

                                    if result == "applied":
                                        if (
                                            existing_contract
                                            and existing_contract.status == CapitatedContract.STATUS_ACTIVE
                                        ):
                                            existing_contract.status = CapitatedContract.STATUS_EXPIRED
                                            existing_contract.terminated_at = datetime.now(timezone.utc)
                                            existing_contract.termination_reason = (
                                                f"Quiebre de continuidad al cargar el mes {coverage_month.strftime('%Y-%m')}."
                                            )
                                            await self._db.flush()

                                        contract = await self._create_new_contract(
                                            company, product, person, plan_version, coverage_month, age  # type: ignore[arg-type]
                                        )
                                        continuity_break = True
                                else:
                                    # Continuidad normal
                                    if existing_contract:
                                        if plan_version.max_renewal_age and age is not None:
                                            if age > plan_version.max_renewal_age:
                                                result = "rejected"
                                                rejection_code = CapitatedRejectionCodes.PERSON_AGE_INVALID
                                                rejection_detail = "Excede edad maxima para renovacion"

                                        if result == "applied":
                                            existing_contract.status = CapitatedContract.STATUS_ACTIVE
                                            existing_contract.valid_until = _end_of_month(coverage_month)
                                            await self._db.flush()
                                            contract = existing_contract
                                    else:
                                        if plan_version.max_entry_age and age is not None:
                                            if age > plan_version.max_entry_age:
                                                result = "rejected"
                                                rejection_code = CapitatedRejectionCodes.PERSON_AGE_INVALID
                                                rejection_detail = "Excede edad maxima para contratacion"

                                        if result == "applied":
                                            contract = await self._create_new_contract(
                                                company, product, person, plan_version, coverage_month, age  # type: ignore[arg-type]
                                            )

                                if continuity_break:
                                    rejection_code = CapitatedRejectionCodes.CONTINUITY_BREAK
                                    rejection_detail = (
                                        f"Se abrio un nuevo contrato por quiebre de continuidad "
                                        f"(ultimo mes {last_month.strftime('%Y-%m')})."
                                    )
                        else:
                            # Primera vez: nuevo contrato
                            if plan_version.max_entry_age and age is not None:
                                if age > plan_version.max_entry_age:
                                    result = "rejected"
                                    rejection_code = CapitatedRejectionCodes.PERSON_AGE_INVALID
                                    rejection_detail = "Excede edad maxima para contratacion"

                            if result == "applied":
                                contract = await self._create_new_contract(
                                    company, product, person, plan_version, coverage_month, age  # type: ignore[arg-type]
                                )

                        # 3) Crear registro mensual y calcular precio
                        if result == "applied" and contract:
                            price_base: float | None = None
                            price_source: str | None = None

                            if res_iso3 and res_iso3 in country_prices:
                                price_base = country_prices[res_iso3]
                                price_source = "country"
                            else:
                                price_base = float(plan_version.price_1 or 0)
                                price_source = "global"

                            # Recargos por edad
                            matched_rules = []
                            for rule in age_surcharges:
                                if rule.age_from is None or rule.age_to is None:
                                    continue
                                if age is not None and rule.age_from <= age <= rule.age_to:
                                    matched_rules.append(rule)

                            age_rule_id: int | None = None
                            age_percent: float | None = None
                            age_amount: float | None = None
                            price_final = price_base

                            if len(matched_rules) > 1:
                                result = "incongruence"
                                rejection_code = CapitatedRejectionCodes.PERSON_INCONGRUENCE
                                rejection_detail = (
                                    "La edad coincide con mas de un tramo de recargo en la version de plan."
                                )
                            elif len(matched_rules) == 1:
                                rule = matched_rules[0]
                                age_rule_id = rule.id
                                age_percent = float(rule.surcharge_percent or 0)
                                price_final = round(
                                    price_base * (100.0 + age_percent) / 100, 2
                                )
                                age_amount = price_final - price_base

                            if result == "applied":
                                monthly_record = CapitatedMonthlyRecord(
                                    company_id=company.id,
                                    product_id=product.id,
                                    person_id=person.id,
                                    contract_id=contract.id,
                                    coverage_month=coverage_month,
                                    plan_version_id=plan_version.id,
                                    load_batch_id=batch.id,
                                    full_name=full_name,
                                    sex=sex,
                                    age_reported=age,
                                    residence_country_id=res_country.id if res_country else None,
                                    repatriation_country_id=rep_country.id if rep_country else None,
                                    price_base=price_base,
                                    price_source=price_source,
                                    age_surcharge_rule_id=age_rule_id,
                                    age_surcharge_percent=age_percent,
                                    age_surcharge_amount=age_amount,
                                    price_final=price_final,
                                    status=CapitatedMonthlyRecord.STATUS_ACTIVE,
                                )
                                self._db.add(monthly_record)
                                await self._db.flush()

                # Contadores
                if result == "applied":
                    total_applied += 1
                elif result == "rejected":
                    total_rejected += 1
                elif result == "duplicated":
                    total_duplicated += 1
                elif result == "incongruence":
                    total_incongruences += 1

                if rejection_code:
                    errors_by_code[rejection_code] = errors_by_code.get(rejection_code, 0) + 1

                # Logging de items segun flags
                should_log = (
                    (result == "applied" and log_applied)
                    or (result == "rejected" and log_rejected)
                    or (result == "incongruence" and log_incongruence)
                    or (result == "duplicated" and log_duplicated)
                )

                if should_log:
                    item_log = CapitatedBatchItemLog(
                        batch_id=batch.id,
                        sheet_name=sheet_name,
                        row_number=row_idx,
                        product_id=product.id,
                        plan_version_id=plan_version.id,
                        residence_raw=res_raw,
                        residence_code_extracted=res_iso3,
                        repatriation_raw=rep_raw,
                        repatriation_code_extracted=rep_iso3,
                        residence_country_id=res_country.id if res_country else None,
                        repatriation_country_id=rep_country.id if rep_country else None,
                        document_number=document_number,
                        full_name=full_name,
                        sex=sex if sex in ("M", "F") else "",
                        age_reported=age,
                        result=result,
                        rejection_code=rejection_code,
                        rejection_detail=rejection_detail,
                        person_id=person.id if person else None,
                        contract_id=contract.id if contract else None,
                        monthly_record_id=monthly_record.id if monthly_record else None,
                        duplicated_record_id=duplicated_record.id if duplicated_record else None,
                    )
                    self._db.add(item_log)

            await self._db.flush()

        # 4) Actualizar batch con resumen
        batch.status = (
            CapitatedBatchLog.STATUS_PROCESSED
            if total_applied > 0
            else CapitatedBatchLog.STATUS_PROCESSED_ZERO
        )
        batch.processed_at = datetime.now(timezone.utc)
        batch.total_rows = total_rows
        batch.total_applied = total_applied
        batch.total_rejected = total_rejected
        batch.total_duplicated = total_duplicated
        batch.total_incongruences = total_incongruences
        batch.total_plan_errors = 0
        batch.total_rolled_back = 0
        batch.summary_json = json.dumps({"errors_by_code": errors_by_code}, ensure_ascii=False)
        await self._db.flush()

    # ------------------------------------------------------------------
    # buildSheetMetas
    # ------------------------------------------------------------------

    async def _build_sheet_metas(
        self, company: Company, wb: Any
    ) -> tuple[dict[str, dict], list[dict]]:
        """Construye metadatos por hoja: producto, plan_version, columnas, paises, recargos."""
        sheet_metas: dict[str, dict] = {}
        plan_errors: list[dict] = []

        for sheet in wb.worksheets:
            sheet_name = (sheet.title or "").strip()

            # Formato "(id) Nombre"
            m = re.match(r"^\s*\((\d+)\)\s*(.+)$", sheet_name)
            if not m:
                continue

            product_id = int(m.group(1))

            # Buscar producto
            r = await self._db.execute(
                select(Product).where(
                    Product.id == product_id,
                    Product.company_id == company.id,
                    Product.product_type == "plan_capitado",
                )
            )
            product = r.scalar_one_or_none()

            if not product:
                plan_errors.append({
                    "code": CapitatedRejectionCodes.PLAN_INVALID_PRODUCT,
                    "sheet": sheet_name,
                    "product_id": product_id,
                    "message": "No existe un producto plan_capitado valido para esta company en la hoja.",
                })
                continue

            # Buscar version activa
            r = await self._db.execute(
                select(PlanVersion).where(
                    PlanVersion.product_id == product.id,
                    PlanVersion.status == "active",
                )
            )
            plan_version = r.scalar_one_or_none()

            if not plan_version:
                plan_errors.append({
                    "code": CapitatedRejectionCodes.PLAN_NO_ACTIVE_VERSION,
                    "sheet": sheet_name,
                    "product_id": product_id,
                    "message": "El producto no tiene version activa.",
                })
                continue

            # Cargar relaciones (countries, repatriationCountries, ageSurcharges)
            # En SQLAlchemy async, las relaciones selectin se cargan automaticamente
            # pero necesitamos acceder a ellas explicitamente
            countries_list = plan_version.countries
            repatriation_countries_list = plan_version.repatriation_countries
            age_surcharges_list = plan_version.age_surcharges

            # Encabezados (primera fila)
            headers: dict[str, int] = {}
            first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=False))
            if first_row:
                for cell in first_row[0]:
                    value = cell.value
                    if value is None:
                        continue
                    normalized = str(value).strip().upper()
                    if normalized:
                        headers[normalized] = cell.column  # 1-based

            required = ["ID", "NOMBRE", "RESIDENCIA", "NACIONALIDAD", "SEXO", "EDAD"]
            missing = [h for h in required if h not in headers]

            if missing:
                plan_errors.append({
                    "code": CapitatedRejectionCodes.PLAN_STRUCTURE_INVALID,
                    "sheet": sheet_name,
                    "product_id": product_id,
                    "missing_headers": missing,
                    "message": "Faltan encabezados obligatorios en la hoja.",
                })
                continue

            # Mapas de paises permitidos y precios
            res_allowed: dict[str, bool] = {}
            country_prices_map: dict[str, float] = {}

            # Obtener precios del pivot table
            pivot_prices = await self._db.execute(
                select(
                    plan_version_countries.c.country_id,
                    plan_version_countries.c.price,
                ).where(plan_version_countries.c.plan_version_id == plan_version.id)
            )
            pivot_by_country_id: dict[int, float | None] = {}
            for pivot_row in pivot_prices.all():
                pivot_by_country_id[pivot_row[0]] = float(pivot_row[1]) if pivot_row[1] is not None else None

            for c in countries_list:
                iso3 = (c.iso3 or "").upper()
                res_allowed[iso3] = True
                if c.id in pivot_by_country_id and pivot_by_country_id[c.id] is not None:
                    country_prices_map[iso3] = pivot_by_country_id[c.id]  # type: ignore[assignment]

            rep_allowed: dict[str, bool] = {}
            for c in repatriation_countries_list:
                rep_allowed[(c.iso3 or "").upper()] = True

            sheet_metas[sheet_name] = {
                "sheet": sheet,
                "product": product,
                "plan_version": plan_version,
                "columns": {h: headers[h] for h in required},
                "residence_iso3": res_allowed,
                "repatriation_iso3": rep_allowed,
                "country_prices": country_prices_map,
                "age_surcharges": age_surcharges_list,
            }

        return sheet_metas, plan_errors

    # ------------------------------------------------------------------
    # createNewContract
    # ------------------------------------------------------------------

    async def _create_new_contract(
        self,
        company: Company,
        product: Product,
        person: CapitatedProductInsured,
        plan_version: PlanVersion,
        coverage_month: date,
        age: int,
    ) -> CapitatedContract:
        """Crea un nuevo contrato."""
        entry_date = coverage_month.replace(day=1)
        valid_until = _end_of_month(coverage_month)

        contract = CapitatedContract(
            company_id=company.id,
            product_id=product.id,
            person_id=person.id,
            status=CapitatedContract.STATUS_ACTIVE,
            entry_date=entry_date,
            valid_until=valid_until,
            entry_age=age,
            wtime_suicide_ends_at=(
                entry_date + timedelta(days=int(plan_version.wtime_suicide))
                if plan_version.wtime_suicide
                else None
            ),
            wtime_preexisting_conditions_ends_at=(
                entry_date + timedelta(days=int(plan_version.wtime_preexisting_conditions))
                if plan_version.wtime_preexisting_conditions
                else None
            ),
            wtime_accident_ends_at=(
                entry_date + timedelta(days=int(plan_version.wtime_accident))
                if plan_version.wtime_accident
                else None
            ),
        )
        self._db.add(contract)
        await self._db.flush()
        return contract

    # ------------------------------------------------------------------
    # Rollback methods
    # ------------------------------------------------------------------

    async def can_rollback_batch(self, batch: CapitatedBatchLog) -> bool:
        """Indica si un batch puede intentar rollback."""
        if batch.status != CapitatedBatchLog.STATUS_PROCESSED:
            return False

        r = await self._db.execute(
            select(CapitatedMonthlyRecord.id)
            .where(
                CapitatedMonthlyRecord.company_id == batch.company_id,
                CapitatedMonthlyRecord.load_batch_id == batch.id,
                CapitatedMonthlyRecord.status == CapitatedMonthlyRecord.STATUS_ACTIVE,
            )
            .limit(1)
        )
        return r.scalar_one_or_none() is not None

    async def can_rollback_monthly_record(self, record: CapitatedMonthlyRecord) -> bool:
        """Indica si un registro mensual puede ser revertido."""
        if record.status == CapitatedMonthlyRecord.STATUS_ROLLED_BACK:
            return False

        if not record.contract_id or not record.coverage_month:
            return False

        r = await self._db.execute(
            select(CapitatedMonthlyRecord.id)
            .where(
                CapitatedMonthlyRecord.company_id == record.company_id,
                CapitatedMonthlyRecord.contract_id == record.contract_id,
                CapitatedMonthlyRecord.coverage_month > record.coverage_month,
                CapitatedMonthlyRecord.status == CapitatedMonthlyRecord.STATUS_ACTIVE,
            )
            .limit(1)
        )
        return r.scalar_one_or_none() is None

    async def rollback_batch(
        self,
        company: Company,
        batch: CapitatedBatchLog,
        user_id: int,
    ) -> CapitatedBatchLog:
        """Rollback completo de un batch (best effort)."""
        if batch.company_id != company.id:
            raise RuntimeError("El batch no pertenece a la empresa indicada.")

        if not await self.can_rollback_batch(batch):
            raise RuntimeError("El batch no es elegible para rollback.")

        r = await self._db.execute(
            select(CapitatedMonthlyRecord).where(
                CapitatedMonthlyRecord.company_id == company.id,
                CapitatedMonthlyRecord.load_batch_id == batch.id,
            )
        )
        records = r.scalars().all()

        if not records:
            return batch

        for record in records:
            if not await self.can_rollback_monthly_record(record):
                continue
            await self._rollback_monthly_record_inner(company, batch, record, user_id, False)

        await self._recalculate_batch_stats(batch, user_id)
        await self._refresh_contracts_status(company)
        await self._db.flush()
        return batch

    async def rollback_monthly_record(
        self,
        company: Company,
        batch: CapitatedBatchLog,
        record: CapitatedMonthlyRecord,
        user_id: int,
        recalculate_batch_stats: bool = True,
    ) -> CapitatedMonthlyRecord:
        """Rollback de un unico registro mensual perteneciente a un batch."""
        await self._rollback_monthly_record_inner(
            company, batch, record, user_id, recalculate_batch_stats
        )
        await self._db.flush()
        return record

    async def _rollback_monthly_record_inner(
        self,
        company: Company,
        batch: CapitatedBatchLog,
        record: CapitatedMonthlyRecord,
        user_id: int,
        recalculate_batch_stats: bool,
    ) -> None:
        """Logica interna de rollback de un registro mensual."""
        if record.company_id != company.id or record.load_batch_id != batch.id:
            raise RuntimeError("El registro mensual no pertenece al batch/empresa indicados.")

        if not await self.can_rollback_monthly_record(record):
            raise RuntimeError("El registro mensual no es elegible para rollback.")

        # Marcar como rolled_back
        record.status = CapitatedMonthlyRecord.STATUS_ROLLED_BACK
        record.rolled_back_at = datetime.now(timezone.utc)
        record.rolled_back_by_user_id = user_id

        # Ajustar contrato
        contract_id = record.contract_id
        if contract_id:
            r = await self._db.execute(
                select(CapitatedContract).where(CapitatedContract.id == contract_id)
            )
            contract = r.scalar_one_or_none()
            if contract:
                r2 = await self._db.execute(
                    select(CapitatedMonthlyRecord)
                    .where(
                        CapitatedMonthlyRecord.company_id == record.company_id,
                        CapitatedMonthlyRecord.contract_id == contract.id,
                        CapitatedMonthlyRecord.status == CapitatedMonthlyRecord.STATUS_ACTIVE,
                    )
                    .order_by(
                        CapitatedMonthlyRecord.coverage_month.desc(),
                        CapitatedMonthlyRecord.id.desc(),
                    )
                    .limit(1)
                )
                last_active_record = r2.scalar_one_or_none()

                if last_active_record:
                    contract.status = CapitatedContract.STATUS_ACTIVE
                    contract.valid_until = _end_of_month(last_active_record.coverage_month)
                else:
                    await self._rollback_contract(contract, user_id)

        if recalculate_batch_stats:
            await self._recalculate_batch_stats(batch, user_id)
            await self._refresh_contracts_status(company)

        await self._db.flush()

    async def _rollback_contract(self, contract: CapitatedContract, user_id: int) -> None:
        """Rollback de contrato cuando no quedan registros mensuales activos."""
        if contract.status == CapitatedContract.STATUS_ROLLED_BACK:
            return

        contract.status = CapitatedContract.STATUS_ROLLED_BACK
        contract.terminated_at = datetime.now(timezone.utc)
        contract.termination_reason = "Rollback: sin registros mensuales activos."

        # Verificar si quedan otros contratos activos para esta persona/producto
        r = await self._db.execute(
            select(CapitatedContract.id)
            .where(
                CapitatedContract.company_id == contract.company_id,
                CapitatedContract.product_id == contract.product_id,
                CapitatedContract.person_id == contract.person_id,
                CapitatedContract.status.in_([
                    CapitatedContract.STATUS_ACTIVE,
                    CapitatedContract.STATUS_EXPIRED,
                ]),
            )
            .limit(1)
        )
        has_active = r.scalar_one_or_none() is not None

        if not has_active:
            await self._rollback_person(contract.person_id, user_id)

    async def _rollback_person(self, person_id: int, user_id: int) -> None:
        """Rollback de ficha de persona cuando no quedan contratos activos."""
        if not person_id:
            return

        r = await self._db.execute(
            select(CapitatedProductInsured).where(CapitatedProductInsured.id == person_id)
        )
        person = r.scalar_one_or_none()
        if not person:
            return

        person.status = CapitatedProductInsured.STATUS_ROLLED_BACK
        person.rolled_back_at = datetime.now(timezone.utc)
        person.rolled_back_by_user_id = user_id

    async def _recalculate_batch_stats(
        self, batch: CapitatedBatchLog, user_id: int | None = None
    ) -> None:
        """Recalcula estadisticas del batch en base a registros mensuales."""
        from sqlalchemy import func

        base_where = [
            CapitatedMonthlyRecord.company_id == batch.company_id,
            CapitatedMonthlyRecord.load_batch_id == batch.id,
        ]

        r1 = await self._db.execute(
            select(func.count())
            .select_from(CapitatedMonthlyRecord)
            .where(*base_where, CapitatedMonthlyRecord.status == CapitatedMonthlyRecord.STATUS_ACTIVE)
        )
        total_active = r1.scalar() or 0

        r2 = await self._db.execute(
            select(func.count())
            .select_from(CapitatedMonthlyRecord)
            .where(*base_where, CapitatedMonthlyRecord.status == CapitatedMonthlyRecord.STATUS_ROLLED_BACK)
        )
        total_rolled_back = r2.scalar() or 0

        batch.total_applied = total_active
        batch.total_rolled_back = total_rolled_back

        if total_active == 0 and total_rolled_back > 0:
            batch.status = CapitatedBatchLog.STATUS_ROLLED_BACK
            batch.rolled_back_at = datetime.now(timezone.utc)
            batch.rolled_back_by_user_id = user_id

    async def _refresh_contracts_status(self, company: Company) -> int:
        """
        Recalcula el status de contratos en base a valid_until vs hoy.

        Usa raw SQL con CASE para eficiencia (UPDATE masivo).
        """
        today = date.today().isoformat()

        stmt = (
            update(CapitatedContract)
            .where(
                CapitatedContract.company_id == company.id,
                CapitatedContract.status.in_([
                    CapitatedContract.STATUS_ACTIVE,
                    CapitatedContract.STATUS_EXPIRED,
                ]),
            )
            .values(
                status=text(
                    f"CASE WHEN valid_until >= '{today}' "
                    f"THEN '{CapitatedContract.STATUS_ACTIVE}' "
                    f"ELSE '{CapitatedContract.STATUS_EXPIRED}' END"
                )
            )
        )
        result = await self._db.execute(stmt)
        return result.rowcount  # type: ignore[return-value]

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    async def _normalize_country_to_iso3(self, raw: str | None) -> str | None:
        """Normaliza un string de pais a ISO3."""
        if not raw:
            return None

        raw = raw.strip().upper()
        if not raw:
            return None

        m = re.match(r"^([A-Z]{2,3})", raw)
        if not m:
            return None

        token = m.group(1)

        # Si token >= 3 letras, probar ISO3
        if len(token) >= 3:
            code3 = token[:3]
            r = await self._db.execute(
                select(Country).where(Country.iso3 == code3)
            )
            country = r.scalar_one_or_none()
            if country:
                return country.iso3

        # Probar ISO2
        code2 = token[:2]
        r = await self._db.execute(
            select(Country).where(Country.iso2 == code2)
        )
        country = r.scalar_one_or_none()
        if country:
            return country.iso3

        return None

    async def _find_country_by_iso3(self, iso3: str) -> Country | None:
        r = await self._db.execute(select(Country).where(Country.iso3 == iso3))
        return r.scalar_one_or_none()

    async def _find_person(
        self, company_id: int, product_id: int, document_number: str
    ) -> CapitatedProductInsured | None:
        r = await self._db.execute(
            select(CapitatedProductInsured).where(
                CapitatedProductInsured.company_id == company_id,
                CapitatedProductInsured.product_id == product_id,
                CapitatedProductInsured.document_number == document_number,
                CapitatedProductInsured.status == CapitatedProductInsured.STATUS_ACTIVE,
            )
        )
        return r.scalar_one_or_none()

    async def _find_active_monthly_record(
        self, company_id: int, product_id: int, person_id: int, coverage_month: date
    ) -> CapitatedMonthlyRecord | None:
        r = await self._db.execute(
            select(CapitatedMonthlyRecord).where(
                CapitatedMonthlyRecord.company_id == company_id,
                CapitatedMonthlyRecord.product_id == product_id,
                CapitatedMonthlyRecord.person_id == person_id,
                CapitatedMonthlyRecord.coverage_month == coverage_month,
                CapitatedMonthlyRecord.status == CapitatedMonthlyRecord.STATUS_ACTIVE,
            )
        )
        return r.scalar_one_or_none()

    async def _find_last_active_record(
        self, company_id: int, product_id: int, person_id: int
    ) -> CapitatedMonthlyRecord | None:
        r = await self._db.execute(
            select(CapitatedMonthlyRecord)
            .where(
                CapitatedMonthlyRecord.company_id == company_id,
                CapitatedMonthlyRecord.product_id == product_id,
                CapitatedMonthlyRecord.person_id == person_id,
                CapitatedMonthlyRecord.status == CapitatedMonthlyRecord.STATUS_ACTIVE,
            )
            .order_by(
                CapitatedMonthlyRecord.coverage_month.desc(),
                CapitatedMonthlyRecord.id.desc(),
            )
            .limit(1)
        )
        return r.scalar_one_or_none()

    async def _find_contract_for_record(
        self, contract_id: int
    ) -> CapitatedContract | None:
        r = await self._db.execute(
            select(CapitatedContract).where(
                CapitatedContract.id == contract_id,
                CapitatedContract.status.in_([
                    CapitatedContract.STATUS_ACTIVE,
                    CapitatedContract.STATUS_EXPIRED,
                ]),
            )
        )
        return r.scalar_one_or_none()

    @staticmethod
    def _cell_str(row_data: tuple, col_index: int) -> str:
        """Devuelve el valor de una celda como string trimmeado (col_index es 1-based)."""
        try:
            cell = row_data[col_index - 1]
            value = cell.value
        except (IndexError, AttributeError):
            return ""

        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _is_row_empty(values: dict[str, str]) -> bool:
        return all(v.strip() == "" for v in values.values())

    @staticmethod
    def _normalize_coverage_month(d: date) -> date:
        return d.replace(day=1)

    @staticmethod
    def _sha1_file(path: str) -> str | None:
        try:
            p = Path(path)
            if not p.is_file():
                return None
            h = hashlib.sha1()
            with p.open("rb") as f:
                for chunk in iter(lambda: f.read(8192), b""):
                    h.update(chunk)
            return h.hexdigest()
        except OSError:
            return None


# ------------------------------------------------------------------
# Date helpers (module-level)
# ------------------------------------------------------------------


def _add_month(d: date) -> date:
    """Suma un mes sin overflow (ej: 2024-01-31 -> 2024-02-29)."""
    if d.month == 12:
        return d.replace(year=d.year + 1, month=1, day=1)
    return d.replace(month=d.month + 1, day=1)


def _end_of_month(d: date) -> date:
    """Retorna el ultimo dia del mes de d."""
    next_month = _add_month(d)
    return next_month - timedelta(days=1)
