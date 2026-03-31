"""
Controlador de lotes capitados y reportes mensuales (admin).
Equivale a CapitatedBatchController.php + CapitatedMonthlyReportController.php.

Endpoints Batches:
  GET    /admin/companies/{cid}/capitados/batches                              → index
  GET    /admin/companies/{cid}/capitados/batches/template                     → template
  GET    /admin/companies/{cid}/capitados/batches/{bid}                        → show
  GET    /admin/companies/{cid}/capitados/batches/{bid}/items                  → items
  GET    /admin/companies/{cid}/capitados/batches/{bid}/monthly-records        → monthlyRecords
  POST   /admin/companies/{cid}/capitados/batches/{bid}/rollback              → rollback
  POST   .../batches/{bid}/monthly-records/{rid}/rollback                      → rollbackMonthlyRecord
  POST   /admin/companies/{cid}/capitados/batches/upload                       → upload

Endpoints Reports:
  GET    /admin/companies/{cid}/capitados/reportes/mensuales                   → months
  GET    /admin/companies/{cid}/capitados/reportes/mensuales/{month}/download  → download
"""

from __future__ import annotations

import io
import math
from datetime import UTC, date, datetime

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import case, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.requests.capitated_batch_request import (
    BatchDetailResponse,
    BatchIndexResponse,
    BatchItemOut,
    BatchItemsResponse,
    BatchOut,
    MonthlyRecordBatchOut,
    MonthlyRecordsBatchResponse,
    MonthlyReportMonthsResponse,
    MonthSummaryOut,
    PaginationMeta,
    RollbackResponse,
)
from common.database import get_db
from common.middleware.permission import require_permission
from common.models.capitated_batch_item_log import CapitatedBatchItemLog
from common.models.capitated_batch_log import CapitatedBatchLog
from common.models.capitated_monthly_record import CapitatedMonthlyRecord
from common.models.company import Company
from common.models.product import Product
from common.models.user import User
from common.services.capitated.capitated_batch_processor import CapitatedBatchProcessor

router = APIRouter(
    prefix="/admin/companies/{company_id}/capitados",
    tags=["admin:capitated-batches"],
)

_PERMISSION = "admin.companies.manage"


# ─────────────────────────── Helpers ─────────────────────────────────────────


async def _get_company(cid: int, db: AsyncSession) -> Company:
    r = await db.execute(select(Company).where(Company.id == cid))
    c = r.scalar_one_or_none()
    if c is None:
        raise HTTPException(status_code=404, detail="Empresa no encontrada.")
    return c


async def _get_batch(cid: int, bid: int, db: AsyncSession) -> CapitatedBatchLog:
    r = await db.execute(
        select(CapitatedBatchLog).where(
            CapitatedBatchLog.id == bid,
            CapitatedBatchLog.company_id == cid,
        )
    )
    b = r.scalar_one_or_none()
    if b is None:
        raise HTTPException(status_code=404, detail="Lote no encontrado.")
    return b


def _batch_out(b: CapitatedBatchLog) -> BatchOut:
    return BatchOut(
        id=b.id,
        company_id=b.company_id,
        coverage_month=str(b.coverage_month) if b.coverage_month else None,
        status=b.status,
        source=b.source,
        original_filename=b.original_filename,
        total_rows=b.total_rows,
        total_applied=b.total_applied,
        total_rejected=b.total_rejected,
        total_duplicated=b.total_duplicated,
        total_incongruences=b.total_incongruences,
        total_plan_errors=b.total_plan_errors,
        total_rolled_back=b.total_rolled_back,
        created_by_user_id=b.created_by_user_id,
        processed_at=str(b.processed_at) if b.processed_at else None,
    )


def _item_out(i: CapitatedBatchItemLog) -> BatchItemOut:
    return BatchItemOut(
        id=i.id,
        batch_id=i.batch_id,
        sheet_name=i.sheet_name,
        row_number=i.row_number,
        product_id=i.product_id,
        document_number=i.document_number,
        full_name=i.full_name,
        sex=i.sex,
        age_reported=i.age_reported,
        result=i.result,
        rejection_code=i.rejection_code,
        rejection_detail=i.rejection_detail,
        residence_raw=i.residence_raw,
        repatriation_raw=i.repatriation_raw,
    )


def _mr_out(mr: CapitatedMonthlyRecord) -> MonthlyRecordBatchOut:
    return MonthlyRecordBatchOut(
        id=mr.id,
        person_id=mr.person_id,
        contract_id=mr.contract_id,
        coverage_month=str(mr.coverage_month) if mr.coverage_month else None,
        full_name=mr.full_name,
        sex=mr.sex,
        age_reported=mr.age_reported,
        price_base=float(mr.price_base) if mr.price_base is not None else None,
        price_final=float(mr.price_final) if mr.price_final is not None else None,
        status=mr.status,
    )


def _make_unique_excel_sheet_title(base_title: str, used_titles: set[str]) -> str:
    """Normaliza, acota a 31 caracteres, evita caracteres inválidos y asegura unicidad."""
    import re as _re

    title = _re.sub(r"[\\/?*\[\]:]", " ", base_title)
    title = _re.sub(r"\s+", " ", title).strip() or "Hoja"
    title = title[:31]

    candidate = title
    i = 2
    while candidate in used_titles:
        suffix = f" ({i})"
        max_len = 31 - len(suffix)
        prefix = title[:max_len] if max_len > 0 else ""
        candidate = prefix + suffix
        i += 1

    used_titles.add(candidate)
    return candidate


def _pagination_meta(total: int, page: int, per_page: int) -> PaginationMeta:
    return PaginationMeta(
        current_page=page,
        last_page=max(1, math.ceil(total / per_page)),
        per_page=per_page,
        total=total,
    )


# ─────────────────────────── Batches ─────────────────────────────────────────


@router.get("/batches", response_model=BatchIndexResponse)
async def batch_index(
    company_id: int,
    status: str = Query(default=""),
    per_page: int = Query(default=15, ge=1, le=100),
    page: int = Query(default=1, ge=1),
    _current_user: User = Depends(require_permission(_PERMISSION)),
    db: AsyncSession = Depends(get_db),
) -> BatchIndexResponse:
    """Lista paginada de lotes de una empresa."""
    await _get_company(company_id, db)

    base_q = select(CapitatedBatchLog).where(CapitatedBatchLog.company_id == company_id)
    if status:
        base_q = base_q.where(CapitatedBatchLog.status == status)

    base_q = base_q.order_by(CapitatedBatchLog.id.desc())

    total = (await db.execute(select(func.count()).select_from(base_q.subquery()))).scalar() or 0
    items = list(
        (await db.execute(base_q.offset((page - 1) * per_page).limit(per_page))).scalars().all()
    )

    return BatchIndexResponse(
        data=[_batch_out(b) for b in items],
        meta=_pagination_meta(total, page, per_page),
    )


@router.get("/batches/template")
async def batch_template(
    company_id: int,
    _current_user: User = Depends(require_permission(_PERMISSION)),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """Descarga plantilla Excel con una hoja por producto capitado activo."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    company = await _get_company(company_id, db)

    # Buscar productos capitados activos con versión de plan activa
    from common.models.plan_version import PlanVersion

    r = await db.execute(
        select(Product)
        .where(
            Product.company_id == company.id,
            Product.product_type == "capitated",
            Product.status == "active",
        )
        .order_by(Product.id)
    )
    all_products = list(r.scalars().all())

    # Filtrar solo los que tienen versión activa
    products = []
    for p in all_products:
        pv_r = await db.execute(
            select(PlanVersion)
            .where(
                PlanVersion.product_id == p.id,
                PlanVersion.status == PlanVersion.STATUS_ACTIVE,
            )
            .limit(1)
        )
        if pv_r.scalar_one_or_none() is not None:
            products.append(p)

    headers = ["ID", "NOMBRE", "RESIDENCIA", "NACIONALIDAD", "SEXO", "EDAD"]
    column_widths = [10, 38, 18, 18, 19, 12]

    wb = openpyxl.Workbook()
    # Eliminar hoja por defecto
    wb.remove(wb.active)

    if not products:
        ws = wb.create_sheet(title="Sin productos")
        ws.append(headers)
        for idx, w in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = w
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        ws.freeze_panes = "A2"
    else:
        used_titles: set[str] = set()
        for product in products:
            product_name = (product.name or "Producto").strip() or "Producto"
            base_title = f"({product.id}) {product_name}"
            title = _make_unique_excel_sheet_title(base_title, used_titles)

            ws = wb.create_sheet(title=title)
            ws.append(headers)
            for idx, w in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(idx)].width = w
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)
            ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    now_str = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    filename = f"capitados_estructura_company_{company.id}_{now_str}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/batches/upload")
async def batch_upload(
    company_id: int,
    file: UploadFile = File(...),
    coverage_month: str | None = Query(default=None),
    _current_user: User = Depends(require_permission(_PERMISSION)),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Sube un Excel y procesa el lote de capitados."""
    company = await _get_company(company_id, db)

    # Validar extensión
    original = file.filename or ""
    if not original.lower().endswith((".xls", ".xlsx")):
        raise HTTPException(status_code=422, detail="El archivo debe ser .xls o .xlsx.")

    # Mes de cobertura
    if coverage_month:
        try:
            cm = datetime.strptime(coverage_month, "%Y-%m-%d").date()
        except ValueError:
            try:
                cm = datetime.strptime(coverage_month, "%Y-%m").date()
            except ValueError:
                raise HTTPException(status_code=422, detail="Formato de mes inválido.")
    else:
        cm = date.today()

    # Normalizar a primer día del mes
    cm = cm.replace(day=1)

    processor = CapitatedBatchProcessor(db)
    batch = await processor.process_excel(
        company=company,
        uploaded_file=file,
        coverage_month=cm,
        is_any_month_allowed=True,
        cutoff_day=15,
        user_id=_current_user.id,
    )

    return {"batch": _batch_out(batch).model_dump()}


@router.get("/batches/{batch_id}", response_model=BatchDetailResponse)
async def batch_show(
    company_id: int,
    batch_id: int,
    _current_user: User = Depends(require_permission(_PERMISSION)),
    db: AsyncSession = Depends(get_db),
) -> BatchDetailResponse:
    """Detalle de un lote."""
    batch = await _get_batch(company_id, batch_id, db)
    return BatchDetailResponse(data=_batch_out(batch))


@router.get("/batches/{batch_id}/items", response_model=BatchItemsResponse)
async def batch_items(
    company_id: int,
    batch_id: int,
    result: str = Query(default=""),
    sheet: str = Query(default=""),
    per_page: int = Query(default=25, ge=1, le=200),
    page: int = Query(default=1, ge=1),
    _current_user: User = Depends(require_permission(_PERMISSION)),
    db: AsyncSession = Depends(get_db),
) -> BatchItemsResponse:
    """Lista paginada de ítems de un lote con filtros."""
    await _get_batch(company_id, batch_id, db)

    base_q = select(CapitatedBatchItemLog).where(CapitatedBatchItemLog.batch_id == batch_id)
    if result:
        base_q = base_q.where(CapitatedBatchItemLog.result == result)
    if sheet:
        base_q = base_q.where(CapitatedBatchItemLog.sheet_name == sheet)

    base_q = base_q.order_by(CapitatedBatchItemLog.row_number)

    total = (await db.execute(select(func.count()).select_from(base_q.subquery()))).scalar() or 0
    items = list(
        (await db.execute(base_q.offset((page - 1) * per_page).limit(per_page))).scalars().all()
    )

    return BatchItemsResponse(
        data=[_item_out(i) for i in items],
        meta=_pagination_meta(total, page, per_page),
    )


@router.get("/batches/{batch_id}/monthly-records", response_model=MonthlyRecordsBatchResponse)
async def batch_monthly_records(
    company_id: int,
    batch_id: int,
    status: str = Query(default=""),
    product_id: int | None = Query(default=None),
    per_page: int = Query(default=25, ge=1, le=200),
    page: int = Query(default=1, ge=1),
    _current_user: User = Depends(require_permission(_PERMISSION)),
    db: AsyncSession = Depends(get_db),
) -> MonthlyRecordsBatchResponse:
    """Lista monthly records generados por un lote."""
    await _get_batch(company_id, batch_id, db)

    base_q = select(CapitatedMonthlyRecord).where(CapitatedMonthlyRecord.load_batch_id == batch_id)
    if status:
        base_q = base_q.where(CapitatedMonthlyRecord.status == status)
    if product_id is not None:
        base_q = base_q.where(CapitatedMonthlyRecord.product_id == product_id)

    base_q = base_q.order_by(CapitatedMonthlyRecord.id)

    total = (await db.execute(select(func.count()).select_from(base_q.subquery()))).scalar() or 0
    items = list(
        (await db.execute(base_q.offset((page - 1) * per_page).limit(per_page))).scalars().all()
    )

    return MonthlyRecordsBatchResponse(
        data=[_mr_out(mr) for mr in items],
        meta=_pagination_meta(total, page, per_page),
    )


# ─────────────────────────── Rollback ────────────────────────────────────────


@router.post("/batches/{batch_id}/rollback", response_model=RollbackResponse)
async def batch_rollback(
    company_id: int,
    batch_id: int,
    _current_user: User = Depends(require_permission(_PERMISSION)),
    db: AsyncSession = Depends(get_db),
) -> RollbackResponse:
    """Rollback completo de un lote (marca monthly records como rolled_back)."""
    batch = await _get_batch(company_id, batch_id, db)

    if batch.status != CapitatedBatchLog.STATUS_PROCESSED:
        raise HTTPException(status_code=422, detail="Solo se pueden revertir lotes procesados.")

    if batch.rolled_back_at is not None:
        raise HTTPException(status_code=422, detail="El lote ya fue revertido.")

    now = datetime.now(UTC)

    # Marcar monthly records activos como rolled_back
    records_r = await db.execute(
        select(CapitatedMonthlyRecord).where(
            CapitatedMonthlyRecord.load_batch_id == batch_id,
            CapitatedMonthlyRecord.status == CapitatedMonthlyRecord.STATUS_ACTIVE,
        )
    )
    count = 0
    for mr in records_r.scalars().all():
        mr.status = CapitatedMonthlyRecord.STATUS_ROLLED_BACK
        mr.rolled_back_at = now
        mr.rolled_back_by_user_id = _current_user.id
        count += 1

    batch.status = CapitatedBatchLog.STATUS_ROLLED_BACK
    batch.rolled_back_at = now
    batch.rolled_back_by_user_id = _current_user.id
    batch.total_rolled_back = count

    await db.commit()

    return RollbackResponse(message=f"Lote revertido. {count} registros afectados.")


@router.post(
    "/batches/{batch_id}/monthly-records/{record_id}/rollback",
    response_model=RollbackResponse,
)
async def rollback_monthly_record(
    company_id: int,
    batch_id: int,
    record_id: int,
    _current_user: User = Depends(require_permission(_PERMISSION)),
    db: AsyncSession = Depends(get_db),
) -> RollbackResponse:
    """Rollback de un solo monthly record."""
    await _get_batch(company_id, batch_id, db)

    r = await db.execute(
        select(CapitatedMonthlyRecord).where(
            CapitatedMonthlyRecord.id == record_id,
            CapitatedMonthlyRecord.load_batch_id == batch_id,
        )
    )
    mr = r.scalar_one_or_none()
    if mr is None:
        raise HTTPException(status_code=404, detail="Registro mensual no encontrado.")

    if mr.status == CapitatedMonthlyRecord.STATUS_ROLLED_BACK:
        raise HTTPException(status_code=422, detail="El registro ya fue revertido.")

    mr.status = CapitatedMonthlyRecord.STATUS_ROLLED_BACK
    mr.rolled_back_at = datetime.now(UTC)
    mr.rolled_back_by_user_id = _current_user.id

    await db.commit()

    return RollbackResponse(message="Registro mensual revertido correctamente.")


# ─────────────────────────── Monthly Reports ─────────────────────────────────


@router.get("/reportes/mensuales", response_model=MonthlyReportMonthsResponse)
async def report_months(
    company_id: int,
    _current_user: User = Depends(require_permission(_PERMISSION)),
    db: AsyncSession = Depends(get_db),
) -> MonthlyReportMonthsResponse:
    """Lista meses disponibles con conteos y totales."""
    await _get_company(company_id, db)

    is_active = CapitatedMonthlyRecord.status == CapitatedMonthlyRecord.STATUS_ACTIVE
    r = await db.execute(
        select(
            CapitatedMonthlyRecord.coverage_month,
            func.sum(case((is_active, 1), else_=0)).label("active_count"),
            func.sum(case((is_active, CapitatedMonthlyRecord.price_final), else_=0)).label(
                "active_total"
            ),
        )
        .where(CapitatedMonthlyRecord.company_id == company_id)
        .group_by(CapitatedMonthlyRecord.coverage_month)
        .order_by(CapitatedMonthlyRecord.coverage_month.desc())
    )

    months = []
    for row in r.all():
        months.append(
            MonthSummaryOut(
                month=str(row[0]) if row[0] else "",
                active_count=row[1] or 0,
                active_total=float(row[2]) if row[2] is not None else None,
            )
        )

    return MonthlyReportMonthsResponse(months=months)


@router.get("/reportes/mensuales/{month}/download")
async def report_download(
    company_id: int,
    month: str,
    _current_user: User = Depends(require_permission(_PERMISSION)),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """Descarga el reporte mensual en Excel."""
    import openpyxl

    await _get_company(company_id, db)

    try:
        coverage_month = datetime.strptime(month, "%Y-%m-%d").date().replace(day=1)
    except ValueError:
        try:
            coverage_month = datetime.strptime(month, "%Y-%m").date().replace(day=1)
        except ValueError:
            raise HTTPException(status_code=422, detail="Formato de mes inválido.")

    r = await db.execute(
        select(CapitatedMonthlyRecord)
        .options(
            selectinload(CapitatedMonthlyRecord.product),
            selectinload(CapitatedMonthlyRecord.residence_country),
            selectinload(CapitatedMonthlyRecord.repatriation_country),
            selectinload(CapitatedMonthlyRecord.contract),
        )
        .where(
            CapitatedMonthlyRecord.company_id == company_id,
            CapitatedMonthlyRecord.coverage_month == coverage_month,
        )
        .order_by(CapitatedMonthlyRecord.product_id, CapitatedMonthlyRecord.id)
    )
    records = list(r.scalars().all())

    # Agrupar por product_id|status
    groups: dict[str, list] = {}
    for rec in records:
        key = f"{rec.product_id}|{rec.status}"
        groups.setdefault(key, []).append(rec)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    headers = [
        "ID",
        "# Contrato",
        "Mes",
        "ID Persona",
        "Persona",
        "Genero",
        "Edad reportada",
        "Residencia ISO3",
        "Residencia ISO2",
        "Residencia",
        "Repatriacion ISO3",
        "Repatriacion ISO2",
        "Repatriacion",
        "Fuente del precio",
        "Precio base",
        "Recargo por edad",
        "Precio total",
    ]

    for _, rows in groups.items():
        first = rows[0]
        product = first.product
        status = first.status

        raw_name = product.name if product else "Producto"
        # Parse translatable JSON name
        if isinstance(raw_name, str) and raw_name.startswith("{"):
            try:
                import json as _json

                parsed = _json.loads(raw_name)
                product_name = parsed.get("es") or parsed.get("en") or raw_name
            except Exception:
                product_name = raw_name
        else:
            product_name = raw_name
        title = f"({product.id if product else '?'}) {product_name}"
        if status == CapitatedMonthlyRecord.STATUS_ROLLED_BACK:
            title = f"({product.id if product else '?'}) RB - {product_name}"

        ws = wb.create_sheet(title=title[:31])
        ws.append(headers)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        for rec in rows:
            age_surcharge = rec.age_surcharge_percent
            if not age_surcharge:
                age_surcharge = "0.00"

            res_c = rec.residence_country
            rep_c = rec.repatriation_country

            ws.append(
                [
                    rec.id,
                    rec.contract_id,
                    rec.coverage_month.strftime("%m/%Y") if rec.coverage_month else "",
                    rec.person_id,
                    rec.full_name,
                    rec.sex,
                    rec.age_reported,
                    res_c.iso3 if res_c else None,
                    res_c.iso2 if res_c else None,
                    res_c.name if res_c else None,
                    rep_c.iso3 if rep_c else None,
                    rep_c.iso2 if rep_c else None,
                    rep_c.name if rep_c else None,
                    rec.price_source,
                    float(rec.price_base) if rec.price_base is not None else None,
                    age_surcharge,
                    float(rec.price_final) if rec.price_final is not None else None,
                ]
            )

    if not wb.sheetnames:
        ws = wb.create_sheet(title="Sin datos")
        ws.append(headers)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"capitados_reporte_{coverage_month.strftime('%Y_%m')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
